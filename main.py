from __future__ import annotations

import re
import json
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

OUTPUT_FILE = "DATAE_APOYOS_2025_INFORME.xlsx"
TEMPLATE_FILE = "dashboard_apoyos_2025.html"
SHEET_ORDER = [
    "RESUMEN_GENERAL",
    "SAN_JOAQUIN",
    "VITACURA",
    "RUT_SIN_CAMPUS",
    "REPORTE_CALIDAD",
    "RUT_INVALIDO_CUARENTENA",
]


def norm_text(value: Any) -> str:
    if value is None:
        return ""
    txt = str(value).strip()
    return "" if txt.lower() in {"", "nan", "none", "null"} else txt


def normalize_rut(value: Any) -> str:
    txt = norm_text(value).upper().replace(".", "")
    txt = re.sub(r"\s+", "", txt)
    txt = re.sub(r"[^0-9K-]", "", txt)
    if not txt:
        return ""

    if "-" in txt:
        body, dv = txt.rsplit("-", 1)
        body = re.sub(r"[^0-9]", "", body)
        dv = re.sub(r"[^0-9K]", "", dv)
        return f"{body}-{dv}" if body and dv else ""

    compact = re.sub(r"[^0-9K]", "", txt)
    if len(compact) >= 2 and re.fullmatch(r"[0-9]+[0-9K]", compact):
        return f"{compact[:-1]}-{compact[-1]}"
    return ""


def expected_dv(body: str) -> str:
    factors = [2, 3, 4, 5, 6, 7]
    total = sum(int(d) * factors[i % len(factors)] for i, d in enumerate(reversed(body)))
    mod = 11 - (total % 11)
    if mod == 11:
        return "0"
    if mod == 10:
        return "K"
    return str(mod)


def validate_rut(rut: str) -> tuple[bool, str]:
    if not rut:
        return False, "RUT vacío"
    if not re.fullmatch(r"\d{7,8}-[0-9K]", rut):
        return False, "Formato inválido (########-X, cuerpo 7/8 dígitos)"
    body, dv = rut.split("-")
    if expected_dv(body) != dv:
        return False, "DV inválido"
    return True, ""


def norm_col(col: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(col).lower())).strip()


def find_col(df: pd.DataFrame, *needles: str) -> str | None:
    cols = {norm_col(c): c for c in df.columns}
    for needle in needles:
        n = norm_col(needle)
        if n in cols:
            return cols[n]
    for c in df.columns:
        n = norm_col(c)
        if any(norm_col(needle) in n for needle in needles):
            return c
    return None


def parse_int(value: Any, default: int = 0) -> int:
    txt = norm_text(value)
    if not txt:
        return default
    m = re.search(r"-?\d+", txt)
    return int(m.group()) if m else default


def campus_from_text(value: Any) -> str | None:
    t = norm_text(value).lower()
    if not t:
        return None
    if "joa" in t or "san" in t:
        return "SAN_JOAQUIN"
    if "vit" in t:
        return "VITACURA"
    return None


def mark_count(n: int) -> str:
    if n <= 0:
        return ""
    return "X" if n == 1 else str(n)


@dataclass
class StudentAgg:
    rut: str
    campus_hints: Counter = field(default_factory=Counter)
    names_by_source: dict[str, Counter] = field(default_factory=lambda: {
        "KATHERINE": Counter(),
        "GLEUDYS": Counter(),
        "PI": Counter(),
        "CIAC": Counter(),
    })
    k_atenciones: int = 0
    g_atenciones: int = 0
    g_mentorias: int = 0
    talleres_pi: int = 0
    talleres_kathy: int = 0
    talleres_gleu: int = 0
    ciac_participaciones: int = 0
    fuentes: set[str] = field(default_factory=set)


BUSINESS_COLS = ["RUT", "Nombre", "Apoyo_Academico_CIAC", "Talleres", "Mentorias", "Atenciones_Individuales"]
TECH_COLS = [
    "fuentes",
    "k_atenciones",
    "g_atenciones",
    "g_mentorias",
    "talleres_pi",
    "talleres_kathy",
    "talleres_gleu",
    "ciac_participaciones",
]


def add_record(
    store: dict[str, StudentAgg],
    invalid_rows: list[dict[str, Any]],
    *,
    rut_raw: Any,
    name: Any = "",
    campus_hint: Any = "",
    source_file: str,
    source_sheet: str,
    row_number: int,
    increments: dict[str, int],
    name_source: str | None = None,
) -> str | None:
    rut_normalized = normalize_rut(rut_raw)
    valid, detail = validate_rut(rut_normalized)
    clean_name = norm_text(name)
    if not valid:
        invalid_rows.append(
            {
                "rut_original": norm_text(rut_raw),
                "rut_normalizado": rut_normalized,
                "fuente": source_file,
                "hoja": source_sheet,
                "fila": row_number,
                "nombre_original": clean_name,
                "motivo": detail,
            }
        )
        return None

    if rut_normalized not in store:
        store[rut_normalized] = StudentAgg(rut=rut_normalized)

    st = store[rut_normalized]
    hint = campus_from_text(campus_hint)
    if hint:
        st.campus_hints[hint] += 1
    if clean_name and name_source:
        st.names_by_source[name_source][clean_name] += 1

    for k, v in increments.items():
        setattr(st, k, getattr(st, k) + int(v))
    st.fuentes.add(source_file)
    return rut_normalized


def load_base(path: Path, campus: str) -> pd.DataFrame:
    df = pd.read_excel(path, dtype=str).fillna("")
    rut_col = find_col(df, "Rut", "RUN")
    dv_col = find_col(df, "DV", "Digito Verificador")
    ap1 = find_col(df, "Apellido 1", "Apellido Paterno")
    ap2 = find_col(df, "Apellido 2", "Apellido Materno")
    nom = find_col(df, "Nombres", "Nombre")

    rows = []
    for idx in range(len(df)):
        raw = f"{df.at[idx, rut_col]}-{df.at[idx, dv_col]}" if rut_col and dv_col else (df.at[idx, rut_col] if rut_col else "")
        rut = normalize_rut(raw)
        valid, reason = validate_rut(rut)
        full_name = " ".join([norm_text(df.at[idx, c]) for c in [nom, ap1, ap2] if c]).strip()
        rows.append({"RUT": rut, "Nombre": full_name, "Campus": campus, "Valido": valid, "Motivo": reason})
    return pd.DataFrame(rows)


def choose_name(st: StudentAgg, sj_name: str, vit_name: str) -> str:
    if sj_name:
        return sj_name
    if vit_name:
        return vit_name
    for source in ["KATHERINE", "GLEUDYS", "PI", "CIAC"]:
        if st.names_by_source[source]:
            return st.names_by_source[source].most_common(1)[0][0]
    return ""


def apply_sheet_format(excel_path: Path) -> None:
    wb = load_workbook(excel_path)
    for sheet in SHEET_ORDER:
        ws = wb[sheet]
        ws.freeze_panes = "A2"
        max_col_letter = ws.cell(row=1, column=max(1, ws.max_column)).column_letter
        ws.auto_filter.ref = f"A1:{max_col_letter}{max(1, ws.max_row)}"
    wb.save(excel_path)


def parse_support_count(value: Any) -> int:
    txt = norm_text(value).upper()
    if not txt:
        return 0
    if txt == "X":
        return 1
    if re.fullmatch(r"\d+", txt):
        return int(txt)
    return 0


def locate_excel_input(repo_root: Path) -> Path:
    candidates = [
        repo_root / "data" / OUTPUT_FILE,
        repo_root / OUTPUT_FILE,
        repo_root / "output" / OUTPUT_FILE,
    ]
    for candidate in candidates:
        if candidate.exists():
            print(f"Excel fuente para dashboard: {candidate}")
            return candidate
    raise FileNotFoundError(
        "No se encontró DATAE_APOYOS_2025_INFORME.xlsx en data/, raíz del repo ni output/."
    )


def locate_template(repo_root: Path) -> Path:
    candidates = [repo_root / TEMPLATE_FILE, repo_root / "data" / TEMPLATE_FILE]
    for candidate in candidates:
        if candidate.exists():
            return candidate
    raise FileNotFoundError(
        "No se encontró dashboard_apoyos_2025.html en la raíz del repo ni en data/."
    )


def read_sheet_best_effort(xls: pd.ExcelFile, sheet_name: str, critical: bool = False) -> pd.DataFrame:
    if sheet_name not in xls.sheet_names:
        msg = f"Hoja faltante: {sheet_name}"
        if critical:
            raise ValueError(f"{msg} (crítica)")
        print(f"[WARN] {msg}")
        return pd.DataFrame()
    return pd.read_excel(xls, sheet_name=sheet_name).fillna("")


def summarize_students_for_payload(df: pd.DataFrame, campus_label: str) -> list[dict[str, Any]]:
    if df.empty:
        return []

    rut_col = find_col(df, "RUT", "RUN")
    name_col = find_col(df, "Nombre", "NOMBRE")
    ciac_col = find_col(df, "Apoyo_Academico_CIAC", "CIAC")
    talleres_col = find_col(df, "Talleres")
    mentorias_col = find_col(df, "Mentorias", "Mentorías")
    atenciones_col = find_col(df, "Atenciones_Individuales", "Atenciones")

    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        rut_original = row[rut_col] if rut_col else ""
        rut_norm = normalize_rut(rut_original)
        rut_valido, _ = validate_rut(rut_norm)

        ciac = parse_support_count(row[ciac_col]) if ciac_col else 0
        talleres = parse_support_count(row[talleres_col]) if talleres_col else 0
        mentorias = parse_support_count(row[mentorias_col]) if mentorias_col else 0
        atenciones = parse_support_count(row[atenciones_col]) if atenciones_col else 0

        rows.append({
            "RUT_NORM": rut_norm,
            "RUT_VALIDO": bool(rut_valido),
            "RUT_ORIGINAL_SI_INVALIDO": norm_text(rut_original) if not rut_valido else "",
            "NOMBRE_NORMALIZADO": norm_text(row[name_col]).upper() if name_col else "",
            "CAMPUS": campus_label,
            "Apoyo_Academico_CIAC": ciac,
            "Talleres": talleres,
            "Mentorias": mentorias,
            "Atenciones_Individuales": atenciones,
            "TOTAL_PARTICIPACIONES": ciac + talleres + mentorias + atenciones,
        })
    return rows


def build_resumen_structure(sj: dict[str, int], vit: dict[str, int], total: dict[str, int]) -> dict[str, dict[str, int]]:
    return {
        "Estudiantes_unicos": {"San_Joaquin": sj["Estudiantes_unicos"], "Vitacura": vit["Estudiantes_unicos"], "Total": total["Estudiantes_unicos"]},
        "Apoyo_CIAC": {"San_Joaquin": sj["Apoyo_CIAC"], "Vitacura": vit["Apoyo_CIAC"], "Total": total["Apoyo_CIAC"]},
        "Talleres": {"San_Joaquin": sj["Talleres"], "Vitacura": vit["Talleres"], "Total": total["Talleres"]},
        "Mentorias": {"San_Joaquin": sj["Mentorias"], "Vitacura": vit["Mentorias"], "Total": total["Mentorias"]},
        "Atenciones_Individuales": {
            "San_Joaquin": sj["Atenciones_Individuales"],
            "Vitacura": vit["Atenciones_Individuales"],
            "Total": total["Atenciones_Individuales"],
        },
    }


def campus_metrics(students: list[dict[str, Any]]) -> dict[str, int]:
    unique_ids = {
        s["RUT_NORM"] if s["RUT_VALIDO"] else f"INVALID:{s['RUT_ORIGINAL_SI_INVALIDO']}"
        for s in students
    }
    return {
        "Estudiantes_unicos": len(unique_ids),
        "Apoyo_CIAC": int(sum(s["Apoyo_Academico_CIAC"] for s in students)),
        "Talleres": int(sum(s["Talleres"] for s in students)),
        "Mentorias": int(sum(s["Mentorias"] for s in students)),
        "Atenciones_Individuales": int(sum(s["Atenciones_Individuales"] for s in students)),
    }


def extract_resumen_general(resumen_df: pd.DataFrame) -> dict[str, dict[str, int]]:
    if resumen_df.empty:
        return build_resumen_structure(
            {"Estudiantes_unicos": 0, "Apoyo_CIAC": 0, "Talleres": 0, "Mentorias": 0, "Atenciones_Individuales": 0},
            {"Estudiantes_unicos": 0, "Apoyo_CIAC": 0, "Talleres": 0, "Mentorias": 0, "Atenciones_Individuales": 0},
            {"Estudiantes_unicos": 0, "Apoyo_CIAC": 0, "Talleres": 0, "Mentorias": 0, "Atenciones_Individuales": 0},
        )

    campus_col = find_col(resumen_df, "Campus")
    row_map = {norm_text(row[campus_col]).upper(): row for _, row in resumen_df.iterrows()} if campus_col else {}

    def metric(row: Any, *aliases: str) -> int:
        if row is None:
            return 0
        for alias in aliases:
            col = find_col(resumen_df, alias)
            if col:
                return parse_int(row[col], 0)
        return 0

    sj_row = row_map.get("SAN_JOAQUIN")
    vit_row = row_map.get("VITACURA")
    total_row = row_map.get("TOTAL")

    sj = {
        "Estudiantes_unicos": metric(sj_row, "Estudiantes_unicos"),
        "Apoyo_CIAC": metric(sj_row, "CIAC_total", "Apoyo_CIAC"),
        "Talleres": metric(sj_row, "Talleres_total", "Talleres"),
        "Mentorias": metric(sj_row, "Mentorias_total", "Mentorias"),
        "Atenciones_Individuales": metric(sj_row, "Atenciones_total", "Atenciones_Individuales"),
    }
    vit = {
        "Estudiantes_unicos": metric(vit_row, "Estudiantes_unicos"),
        "Apoyo_CIAC": metric(vit_row, "CIAC_total", "Apoyo_CIAC"),
        "Talleres": metric(vit_row, "Talleres_total", "Talleres"),
        "Mentorias": metric(vit_row, "Mentorias_total", "Mentorias"),
        "Atenciones_Individuales": metric(vit_row, "Atenciones_total", "Atenciones_Individuales"),
    }
    total = {
        "Estudiantes_unicos": metric(total_row, "Estudiantes_unicos") or (sj["Estudiantes_unicos"] + vit["Estudiantes_unicos"]),
        "Apoyo_CIAC": metric(total_row, "CIAC_total", "Apoyo_CIAC") or (sj["Apoyo_CIAC"] + vit["Apoyo_CIAC"]),
        "Talleres": metric(total_row, "Talleres_total", "Talleres") or (sj["Talleres"] + vit["Talleres"]),
        "Mentorias": metric(total_row, "Mentorias_total", "Mentorias") or (sj["Mentorias"] + vit["Mentorias"]),
        "Atenciones_Individuales": metric(total_row, "Atenciones_total", "Atenciones_Individuales") or (sj["Atenciones_Individuales"] + vit["Atenciones_Individuales"]),
    }
    return build_resumen_structure(sj, vit, total)


def load_excel_build_payload(excel_path: Path) -> dict[str, Any]:
    xls = pd.ExcelFile(excel_path)
    total_rows_all_sheets = 0
    for sheet in xls.sheet_names:
        total_rows_all_sheets += len(pd.read_excel(xls, sheet_name=sheet))

    resumen_df = read_sheet_best_effort(xls, "RESUMEN_GENERAL", critical=False)
    sj_df = read_sheet_best_effort(xls, "SAN_JOAQUIN", critical=True)
    vit_df = read_sheet_best_effort(xls, "VITACURA", critical=True)
    calidad_df = read_sheet_best_effort(xls, "REPORTE_CALIDAD", critical=False)

    students = summarize_students_for_payload(sj_df, "San Joaquín") + summarize_students_for_payload(vit_df, "Vitacura")

    sj_calc = campus_metrics([s for s in students if s["CAMPUS"] == "San Joaquín"])
    vit_calc = campus_metrics([s for s in students if s["CAMPUS"] == "Vitacura"])
    total_calc = {
        k: sj_calc[k] + vit_calc[k] for k in ["Apoyo_CIAC", "Talleres", "Mentorias", "Atenciones_Individuales"]
    }
    total_calc["Estudiantes_unicos"] = sj_calc["Estudiantes_unicos"] + vit_calc["Estudiantes_unicos"]

    resumen_general = extract_resumen_general(resumen_df)
    totales_calculados = build_resumen_structure(sj_calc, vit_calc, total_calc)

    print("Comparación RESUMEN_GENERAL vs totales calculados:")
    for metric in ["Estudiantes_unicos", "Apoyo_CIAC", "Talleres", "Mentorias", "Atenciones_Individuales"]:
        for campus in ["San_Joaquin", "Vitacura", "Total"]:
            expected = resumen_general[metric][campus]
            calculated = totales_calculados[metric][campus]
            if expected != calculated:
                print(f"  [DIFF] {metric} - {campus}: resumen={expected}, calculado={calculated}")

    estudiantes_con_apoyo = sum(1 for s in students if s["TOTAL_PARTICIPACIONES"] > 0)
    estudiantes_unicos_total = total_calc["Estudiantes_unicos"]
    pct_apoyo = round((estudiantes_con_apoyo / estudiantes_unicos_total) * 100, 2) if estudiantes_unicos_total else 0.0

    payload: dict[str, Any] = {
        "meta": {
            "generated_at_iso": datetime.utcnow().isoformat(timespec="seconds"),
            "total_rows_all_sheets": int(total_rows_all_sheets),
        },
        "resumen_general": resumen_general,
        "totales_calculados": totales_calculados,
        "kpis": {
            "estudiantes_unicos_total": estudiantes_unicos_total,
            "apoyo_ciac_total": total_calc["Apoyo_CIAC"],
            "talleres_total": total_calc["Talleres"],
            "mentorias_total": total_calc["Mentorias"],
            "atenciones_total": total_calc["Atenciones_Individuales"],
            "pct_estudiantes_con_apoyo": pct_apoyo,
        },
        "campus_totals": {
            "San Joaquín": sj_calc,
            "Vitacura": vit_calc,
        },
        "students": students,
    }
    if not calidad_df.empty:
        payload["reporte_calidad"] = calidad_df.to_dict(orient="records")
    return payload


def render_dashboard(template_path: Path, payload: dict[str, Any], out_path: Path) -> None:
    html = template_path.read_text(encoding="utf-8")
    replacement = f"const DATOS = {json.dumps(payload, ensure_ascii=False)};"
    pattern = re.compile(r"const\s+DATOS\s*=\s*\{.*?\};", re.DOTALL)
    if not pattern.search(html):
        raise ValueError("No se encontró el bloque 'const DATOS = {...};' en la plantilla.")
    rendered = pattern.sub(replacement, html, count=1)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(rendered, encoding="utf-8")


def run_pipeline(repo_root: Path) -> None:
    data_dir = repo_root / "data"
    output_dir = repo_root / "output"
    debug_dir = output_dir / "_debug"
    output_dir.mkdir(exist_ok=True)
    debug_dir.mkdir(exist_ok=True)

    base_sj = load_base(data_dir / "SAN JOAQUIN APOYOS 2025.xlsx", "SAN_JOAQUIN")
    base_vit = load_base(data_dir / "VITACURA APOYOS 2025.xlsx", "VITACURA")
    base_sj_valid = base_sj[base_sj["Valido"]].drop_duplicates("RUT")
    base_vit_valid = base_vit[base_vit["Valido"]].drop_duplicates("RUT")
    sj_map = dict(zip(base_sj_valid["RUT"], base_sj_valid["Nombre"]))
    vit_map = dict(zip(base_vit_valid["RUT"], base_vit_valid["Nombre"]))

    invalid_rows: list[dict[str, Any]] = []
    students: dict[str, StudentAgg] = {}
    mentorias_valid_source: set[str] = set()
    g_atenciones_valid_source: set[str] = set()

    pi_s1 = pd.read_excel(data_dir / "Asistencia Talleres Proyecto Inicial Santiago 2025-1.xlsx", dtype=str).fillna("")
    rut_col = find_col(pi_s1, "RUT")
    campus_col = find_col(pi_s1, "campus")
    name_col = find_col(pi_s1, "Nombre2", "Nombre")
    for i in range(len(pi_s1)):
        add_record(students, invalid_rows, rut_raw=pi_s1.at[i, rut_col], name=pi_s1.at[i, name_col] if name_col else "", campus_hint=pi_s1.at[i, campus_col] if campus_col else "", source_file="PI_S1", source_sheet="Sheet1", row_number=i + 2, increments={"talleres_pi": 1}, name_source="PI")

    pi_s2 = pd.read_excel(data_dir / "AsistenciaTalleres Proyecto Inicial 2025-2.xlsx", dtype=str).fillna("")
    rut_col = find_col(pi_s2, "RUT")
    campus_col = find_col(pi_s2, "campus")
    name_col = find_col(pi_s2, "Nombre2", "Nombre")
    for i in range(len(pi_s2)):
        add_record(students, invalid_rows, rut_raw=pi_s2.at[i, rut_col], name=pi_s2.at[i, name_col] if name_col else "", campus_hint=pi_s2.at[i, campus_col] if campus_col else "", source_file="PI_S2", source_sheet="Sheet1", row_number=i + 2, increments={"talleres_pi": 1}, name_source="PI")

    ciac_sj = pd.read_excel(data_dir / "CIAC San Joaquín - Registro participación estudiantes.xlsx", sheet_name="Asistencia 2S 2025", dtype=str).fillna("")
    rut_col = find_col(ciac_sj, "RUN")
    for i in range(len(ciac_sj)):
        add_record(students, invalid_rows, rut_raw=ciac_sj.at[i, rut_col], source_file="CIAC_SJ", source_sheet="Asistencia 2S 2025", row_number=i + 2, increments={"ciac_participaciones": 1}, name_source="CIAC")

    ciac_vit = pd.read_excel(data_dir / "CIAC Vitacura - Registro participación estudiantes.xlsx", sheet_name="Registro de participación", dtype=str).fillna("")
    run_col = find_col(ciac_vit, "RUN (sin puntos ni digito verificador)", "RUN")
    dv_col = find_col(ciac_vit, "Dígito Verificador", "DV")
    for i in range(len(ciac_vit)):
        rut_raw = f"{ciac_vit.at[i, run_col]}-{ciac_vit.at[i, dv_col]}" if dv_col else ciac_vit.at[i, run_col]
        add_record(students, invalid_rows, rut_raw=rut_raw, source_file="CIAC_VIT", source_sheet="Registro de participación", row_number=i + 2, increments={"ciac_participaciones": 1}, name_source="CIAC")

    kath_at = pd.read_excel(data_dir / "Katherine - AtencionesyTalleresPsi2025.xlsx", sheet_name="AtenciónIndividual2025", dtype=str).fillna("")
    rut_col = find_col(kath_at, "Rut")
    dv_col = find_col(kath_at, "DV")
    name_cols = [find_col(kath_at, "Nombres"), find_col(kath_at, "Apellido 1"), find_col(kath_at, "Apellido 2")]
    total_col = find_col(kath_at, "Total de sesiones")
    for i in range(len(kath_at)):
        rut_raw = f"{kath_at.at[i, rut_col]}-{kath_at.at[i, dv_col]}" if dv_col else kath_at.at[i, rut_col]
        name = " ".join(norm_text(kath_at.at[i, c]) for c in name_cols if c).strip()
        n = max(0, parse_int(kath_at.at[i, total_col], 0))
        if n > 0:
            add_record(students, invalid_rows, rut_raw=rut_raw, name=name, source_file="KATH_ATENC", source_sheet="AtenciónIndividual2025", row_number=i + 2, increments={"k_atenciones": n}, name_source="KATHERINE")

    kath_taller = pd.read_excel(data_dir / "Katherine - AtencionesyTalleresPsi2025.xlsx", sheet_name="Talleres2025", dtype=str).fillna("")
    rut_col = find_col(kath_taller, "RUT")
    campus_col = find_col(kath_taller, "Campus")
    name_col = find_col(kath_taller, "Nombre")
    for i in range(len(kath_taller)):
        add_record(students, invalid_rows, rut_raw=kath_taller.at[i, rut_col], name=kath_taller.at[i, name_col] if name_col else "", campus_hint=kath_taller.at[i, campus_col] if campus_col else "", source_file="KATH_TALLER", source_sheet="Talleres2025", row_number=i + 2, increments={"talleres_kathy": 1}, name_source="KATHERINE")

    gleu_at = pd.read_excel(data_dir / "Gleudys - Datos implementación 2025 DATA PACE.xlsx", sheet_name="Atenciones individuales", header=1, dtype=str).fillna("")
    rut_col = find_col(gleu_at, "RUT")
    ap1 = find_col(gleu_at, "APELLIDO P")
    ap2 = find_col(gleu_at, "APELLIDO M")
    nom = find_col(gleu_at, "NOMBRE")
    at_col = find_col(gleu_at, "N° de atenciones", "No de atenciones")
    camp_col = find_col(gleu_at, "Campus")
    for i in range(len(gleu_at)):
        name = " ".join(norm_text(gleu_at.at[i, c]) for c in [nom, ap1, ap2] if c).strip()
        n = max(0, parse_int(gleu_at.at[i, at_col], 0))
        if n > 0:
            rut = add_record(students, invalid_rows, rut_raw=gleu_at.at[i, rut_col], name=name, campus_hint=gleu_at.at[i, camp_col] if camp_col else "", source_file="GLEU_ATENC", source_sheet="Atenciones individuales", row_number=i + 3, increments={"g_atenciones": n}, name_source="GLEUDYS")
            if rut:
                g_atenciones_valid_source.add(rut)

    gleu_ment = pd.read_excel(data_dir / "Gleudys - Datos implementación 2025 DATA PACE.xlsx", sheet_name="Mentorías", header=1, dtype=str).fillna("")
    rut_col = find_col(gleu_ment, "Inscritos")
    name_col = find_col(gleu_ment, "Nombre y apellido")
    for i in range(len(gleu_ment)):
        rut = add_record(students, invalid_rows, rut_raw=gleu_ment.at[i, rut_col], name=gleu_ment.at[i, name_col] if name_col else "", source_file="GLEU_MENT", source_sheet="Mentorías", row_number=i + 3, increments={"g_mentorias": 1}, name_source="GLEUDYS")
        if rut:
            mentorias_valid_source.add(rut)

    all_ruts = set(sj_map) | set(vit_map) | set(students)
    rows = []
    alias_rows = []
    for rut in sorted(all_ruts):
        st = students.get(rut, StudentAgg(rut=rut))
        if rut in sj_map:
            campus = "SAN_JOAQUIN"
        elif rut in vit_map:
            campus = "VITACURA"
        else:
            campus = "RUT_SIN_CAMPUS"

        name = choose_name(st, sj_map.get(rut, ""), vit_map.get(rut, ""))
        if rut in sj_map or rut in vit_map:
            padron_name = sj_map.get(rut, "") or vit_map.get(rut, "")
            for source, names in st.names_by_source.items():
                if not names:
                    continue
                src_name = names.most_common(1)[0][0]
                if norm_text(padron_name).lower() != norm_text(src_name).lower():
                    alias_rows.append({
                        "tipo": "ALIAS_NOMBRE",
                        "rut": rut,
                        "nombre_padron": padron_name,
                        "nombre_fuente": src_name,
                        "fuente": source,
                        "tipo_diferencia": "padron_vs_fuente",
                    })

        talleres_total = st.talleres_pi + st.talleres_kathy + st.talleres_gleu
        atenciones_total = st.k_atenciones + st.g_atenciones
        rows.append({
            "RUT": rut,
            "Nombre": name,
            "Campus": campus,
            "Apoyo_Academico_CIAC": mark_count(st.ciac_participaciones),
            "Talleres": mark_count(talleres_total),
            "Mentorias": mark_count(st.g_mentorias),
            "Atenciones_Individuales": mark_count(atenciones_total),
            "fuentes": "|".join(sorted(st.fuentes)),
            "k_atenciones": st.k_atenciones,
            "g_atenciones": st.g_atenciones,
            "g_mentorias": st.g_mentorias,
            "talleres_pi": st.talleres_pi,
            "talleres_kathy": st.talleres_kathy,
            "talleres_gleu": st.talleres_gleu,
            "ciac_participaciones": st.ciac_participaciones,
        })

    final = pd.DataFrame(rows)
    final_sj = final[final["Campus"] == "SAN_JOAQUIN"][BUSINESS_COLS + TECH_COLS].reset_index(drop=True)
    final_vit = final[final["Campus"] == "VITACURA"][BUSINESS_COLS + TECH_COLS].reset_index(drop=True)
    final_sin = final[final["Campus"] == "RUT_SIN_CAMPUS"][BUSINESS_COLS + TECH_COLS].reset_index(drop=True)

    invalid_df = pd.DataFrame(invalid_rows)
    if invalid_df.empty:
        invalid_df = pd.DataFrame(columns=["rut_original", "rut_normalizado", "fuente", "hoja", "fila", "nombre_original", "motivo"])

    mentorias_reflejadas = set(final.loc[final["g_mentorias"] > 0, "RUT"])
    mentorias_faltantes = sorted(mentorias_valid_source - mentorias_reflejadas)
    g_at_reflejadas = set(final.loc[final["g_atenciones"] > 0, "RUT"])
    g_at_faltantes = sorted(g_atenciones_valid_source - g_at_reflejadas)

    double_source = final[(final["k_atenciones"] > 0) & (final["g_atenciones"] > 0)].copy()
    double_source["atenciones_esperadas"] = double_source["k_atenciones"] + double_source["g_atenciones"]
    double_source["atenciones_marcadas"] = double_source["Atenciones_Individuales"].replace("X", "1").replace("", "0").astype(int)
    atenciones_mismatch = double_source[double_source["atenciones_esperadas"] != double_source["atenciones_marcadas"]]

    ciac_marked_without_source = final[(final["Apoyo_Academico_CIAC"] != "") & (final["ciac_participaciones"] == 0)]["RUT"].tolist()
    ciac_off_by_one = final[(final["ciac_participaciones"] > 0) & (final["Apoyo_Academico_CIAC"] != final["ciac_participaciones"].apply(mark_count))]["RUT"].tolist()

    quality_rows = [
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "mentorias_total_fuente", "valor": len(mentorias_valid_source), "detalle": ""},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "mentorias_total_reflejadas", "valor": len(mentorias_reflejadas), "detalle": ""},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "mentorias_faltantes", "valor": len(mentorias_faltantes), "detalle": "|".join(mentorias_faltantes)},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "g_atenciones_total_fuente", "valor": len(g_atenciones_valid_source), "detalle": ""},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "g_atenciones_total_reflejadas", "valor": len(g_at_reflejadas), "detalle": ""},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "g_atenciones_faltantes", "valor": len(g_at_faltantes), "detalle": "|".join(g_at_faltantes)},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "rut_invalidos_total", "valor": len(invalid_df), "detalle": ""},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "ciac_sin_respaldo_total", "valor": len(ciac_marked_without_source), "detalle": "|".join(ciac_marked_without_source)},
        {"tipo": "AUDITORIA_RESUMEN", "metrica": "ciac_off_by_one_total", "valor": len(ciac_off_by_one), "detalle": "|".join(ciac_off_by_one)},
        {"tipo": "CHECKS", "metrica": "mentorias_faltantes_check", "valor": int(len(mentorias_faltantes) == 0), "detalle": "1=OK"},
        {"tipo": "CHECKS", "metrica": "g_atenciones_faltantes_check", "valor": int(len(g_at_faltantes) == 0), "detalle": "1=OK"},
        {"tipo": "CHECKS", "metrica": "suma_doble_fuente_atenciones_mismatch", "valor": len(atenciones_mismatch), "detalle": ""},
        {"tipo": "CHECKS", "metrica": "off_by_one_detected", "valor": int(len(ciac_off_by_one) > 0), "detalle": ""},
        {"tipo": "CHECKS", "metrica": "ciac_marked_without_source", "valor": len(ciac_marked_without_source), "detalle": "|".join(ciac_marked_without_source)},
    ]
    quality_df = pd.DataFrame(quality_rows + alias_rows)

    resumen_rows = []
    for campus, frame in [("SAN_JOAQUIN", final_sj), ("VITACURA", final_vit), ("TOTAL", final[BUSINESS_COLS + TECH_COLS])]:
        resumen_rows.append({
            "Campus": campus,
            "Estudiantes_unicos": int(frame["RUT"].nunique()),
            "CIAC_total": int(frame["ciac_participaciones"].sum()),
            "Talleres_total": int((frame["talleres_pi"] + frame["talleres_kathy"] + frame["talleres_gleu"]).sum()),
            "Mentorias_total": int(frame["g_mentorias"].sum()),
            "Atenciones_total": int((frame["k_atenciones"] + frame["g_atenciones"]).sum()),
        })
    resumen_df = pd.DataFrame(resumen_rows)

    final_sj.to_csv(debug_dir / "SAN_JOAQUIN.csv", index=False)
    final_vit.to_csv(debug_dir / "VITACURA.csv", index=False)
    final_sin.to_csv(debug_dir / "RUT_SIN_CAMPUS.csv", index=False)
    quality_df.to_csv(debug_dir / "REPORTE_CALIDAD.csv", index=False)
    invalid_df.to_csv(debug_dir / "RUT_INVALIDO_CUARENTENA.csv", index=False)

    out_xlsx = output_dir / OUTPUT_FILE
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        resumen_df.to_excel(writer, sheet_name="RESUMEN_GENERAL", index=False)
        final_sj.to_excel(writer, sheet_name="SAN_JOAQUIN", index=False)
        final_vit.to_excel(writer, sheet_name="VITACURA", index=False)
        final_sin.to_excel(writer, sheet_name="RUT_SIN_CAMPUS", index=False)
        quality_df.to_excel(writer, sheet_name="REPORTE_CALIDAD", index=False)
        invalid_df.to_excel(writer, sheet_name="RUT_INVALIDO_CUARENTENA", index=False)

    apply_sheet_format(out_xlsx)
    print(f"Excel generado: {out_xlsx}")


def main() -> None:
    repo_root = Path(__file__).resolve().parent
    run_pipeline(repo_root)

    excel_path = locate_excel_input(repo_root)
    try:
        template_path = locate_template(repo_root)
    except FileNotFoundError as exc:
        print(f"[WARN] {exc}")
        return

    payload = load_excel_build_payload(excel_path)
    dashboard_output = repo_root / "output" / TEMPLATE_FILE
    render_dashboard(template_path, payload, dashboard_output)
    (repo_root / "dashboard.html").write_text(dashboard_output.read_text(encoding="utf-8"), encoding="utf-8")
    print(f"Dashboard generado: {dashboard_output}")


if __name__ == "__main__":
    main()
